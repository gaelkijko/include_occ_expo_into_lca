'''
Created on 16 juil. 2013

ouvre une boite de dialogue

@author: Lenovo
source: http://tkinter.unpythonic.net/wiki/tkFileDialog
'''

#for import_conc the input file is an excel file with  one ligne header:
#NAICS_2007	NAICS_description	CAS			chem_name	conc (mg/m3)	factor SIC to NAICS 2007	year	couple_code
#NAICS-code		str				******-**-*	str			float			float						int		concatenate(NAICS_2007, "-", "CAS)
#outputs a list: [{ sector1_chem1: [conc1,conc2,conc3...] , sector2_chem2 : [conc1,conc2,conc3...]} ,...,  sector2_chem1 : [...] ...},nb_row of initial fila, nb column initial file]

def import_conc(string_file_to_open,generic_file_name,default_extension,possible_extensions):
    import Tkinter, tkFileDialog
    from openpyxl import load_workbook
    root= Tkinter.Tk()
    root.withdraw()
    file_opt = options = {}
    options['defaultextension'] = default_extension
    options['filetypes'] = possible_extensions
    options['initialdir'] = 'D:\\programmation\workspace\eclipse\maitrise'
    options['initialfile'] = generic_file_name + '.xlsx'
    options['parent'] = root
    options['title'] = string_file_to_open
    file_path = tkFileDialog.askopenfilename(**file_opt)

    wb = load_workbook(filename = file_path, use_iterators = True, data_only=True)
    ws = wb.get_active_sheet()
    i=0
    
    dict_data = {}
    headers = ["NAICS","NAICS_description","CAS","CAS_name","concentration","weight","year","couple_id"]
    first = 0
    for row in ws.iter_rows():
        count = 0
        if first == 0:
            first = 1
            continue
        else:
            data = {}
            for cell in row:
                data[headers[count]] = cell.value
                count += 1
            i+=1
        data["NAICS"] = int(data["NAICS"])
        data["year"] = int (data["year"])
        if dict_data.setdefault(str(data["NAICS"]) + "_" +data["CAS"], None) == None:
            dict_data[str(data["NAICS"]) + "_" +data["CAS"]] = [data]
        else:
            dict_data[str(data["NAICS"]) + "_" +data["CAS"]].append(data)
    data=[dict_data,ws.get_highest_row(),ws.get_highest_column()]
    return data;

#for import_hours the input file is an excel file with  one ligne header:
#NAICS_2007	total blue collar hours worked
#NAICS-code	int
#output a list: [{sector1 : [naics_code,hours] , sector2 : [naics_code,hours]},nb_row of initial fila, nb column initial file]

def import_hours(string_file_to_open,generic_file_name,default_extension,possible_extensions):
    import Tkinter, tkFileDialog
    from openpyxl import load_workbook
    root= Tkinter.Tk()
    root.withdraw()
    file_opt = options = {}
    options['defaultextension'] = default_extension
    options['filetypes'] = possible_extensions
    options['initialdir'] = 'D:\\programmation\workspace\eclipse\maitrise'
    options['initialfile'] = generic_file_name + '.xlsx'
    options['parent'] = root
    options['title'] = string_file_to_open
    file_path = tkFileDialog.askopenfilename(**file_opt)

    wb = load_workbook(filename = file_path, use_iterators = True, data_only=True)
    ws = wb.get_active_sheet()
    i=0
    
    dict_data = {}
    headers = ["NAICS","hours"]
    first = 0
    for row in ws.iter_rows():
        count = 0
        if first == 0:
            first = 1
            continue
        else:
            data = {}
            for cell in row:
                data[headers[count]] = cell.value
                count += 1
            i+=1
        data["NAICS"] = int(data["NAICS"])
        if dict_data.setdefault(str(data["NAICS"]), None) == None:
            dict_data[str(data["NAICS"])] = data
        else:
            print "error: duplicate sector line in hours file"
            quit()
    data=[dict_data,ws.get_highest_row(),ws.get_highest_column()]
    return data;

#for import_sectors the input file is an excel file with  one ligne header:
#"NAICS"	"NAICS_description"	"Digits"										"inhalation_rate" (m3/j)
#int		str					str (length of the NAICS code, should be [3-6]	float
#output a list : [{sector1 : [NAICS_code, NAICS_name, digits, inhalation_rate] , sector2 : [] ...},nb_row of initial fila, nb column initial file]

def import_sectors(string_file_to_open,generic_file_name,default_extension,possible_extensions):
    import Tkinter, tkFileDialog
    from openpyxl import load_workbook
    root= Tkinter.Tk()
    root.withdraw()
    file_opt = options = {}
    options['defaultextension'] = default_extension
    options['filetypes'] = possible_extensions
    options['initialdir'] = 'D:\\programmation\workspace\eclipse\maitrise'
    options['initialfile'] = generic_file_name + '.xlsx'
    options['parent'] = root
    options['title'] = string_file_to_open
    file_path = tkFileDialog.askopenfilename(**file_opt)

    wb = load_workbook(filename = file_path, use_iterators = True, data_only=True)
    ws = wb.get_active_sheet()
    i=0
    
    dict_data = {}
    headers = ["NAICS","NAICS_description","Digits","inhalation_rate"]
    first = 0
    for row in ws.iter_rows():
        count = 0
        if first == 0:
            first = 1
            continue
        else:
            data = {}
            for cell in row:
                data[headers[count]] = cell.value
                count += 1
            i+=1
        data["NAICS"] = int(data["NAICS"])
        data["Digits"] = int(data["Digits"])
        if dict_data.setdefault(str(data["NAICS"]), None) == None:
            dict_data[str(data["NAICS"])] = data
        else:
            print "error: duplicate sector line in sector file"
            quit()
    data=[dict_data,ws.get_highest_row(),ws.get_highest_column()]
    return data;

#for import_conc_by_chem the input file is an excel file with  one ligne header:
#NAICS_2007	NAICS_description	CAS			chem_name	conc (mg/m3)	factor SIC to NAICS 2007	year	couple_code
#NAICS-code		str				******-**-*	str			float			float						int		concatenate(NAICS_2007, "-", "CAS)
#the difference with "import_conc" is that it outputs a list:  [{ chem1 : {sector1: [conc1,conc2,conc3...] , sector2 : [conc1,conc2,conc3...]} , chem2 : {...} ...},nb_row of initial fila, nb column initial file]

def import_conc_by_chem(string_file_to_open,generic_file_name,default_extension,possible_extensions):
    import Tkinter, tkFileDialog
    from openpyxl import load_workbook
    root= Tkinter.Tk()
    root.withdraw()
    file_opt = options = {}
    options['defaultextension'] = default_extension
    options['filetypes'] = possible_extensions
    options['initialdir'] = 'D:\\programmation\workspace\eclipse\maitrise'
    options['initialfile'] = generic_file_name + '.xlsx'
    options['parent'] = root
    options['title'] = string_file_to_open
    file_path = tkFileDialog.askopenfilename(**file_opt)

    wb = load_workbook(filename = file_path, use_iterators = True, data_only=True)
    ws = wb.get_active_sheet()
    i=0
    
    dict_data = {}
    headers = ["NAICS","NAICS_description","CAS","CAS_name","concentration","weight","year","couple_id"]
    first = 0
    for row in ws.iter_rows():
        count = 0
        if first == 0:
            first = 1
            continue
        else:
            data = {}
            for cell in row:
                data[headers[count]] = cell.value
                count += 1
            i+=1
        data["NAICS"] = int(data["NAICS"])
        data["year"] = int (data["year"])
        if dict_data.setdefault(data["CAS"], {}) == {}:
            dict_data[data["CAS"]][data["NAICS"]] = [data]
        else:
            dict_data[data["CAS"]].setdefault(data["NAICS"], [])
            dict_data[data["CAS"]][data["NAICS"]].append(data)
    data=[dict_data,ws.get_highest_row(),ws.get_highest_column()]
    return data;

#for import_ED_50_by_CAS the input file is an excel file with  one ligne header:
#CAS			ED50inh,noncanc-kg.lifetime-1	ED50ing,noncanc-kg.lifetime-1	ED50inh,canc-kg.lifetime-1	ED50ing,canc-kg.lifetime-1	GSD2 cancer			GSD2 non cancer		Flag canc inh	Flag canc ing								Flag non canc inh							Flag non canc ing
#******-**-*	float or "n/a"					float or "n/a"					float or "n/a"				float or "n/a"				float (usetox 26,5)	float (usetox 26,5)					"r" (recommended) or "F" (flagged) or "n/a"	"r" (recommended) or "F" (flagged) or "n/a"	"r" (recommended) or "F" (flagged) or "n/a"	
#outputs alist: [{CAS1 : [CAS,ED_50_inh_non_canc_lifetime_kg....] , CAS2 : [] .....},nb_row of initial fila, nb column initial file]

def import_ED_50_by_CAS(string_file_to_open,generic_file_name,default_extension,possible_extensions):
    import Tkinter, tkFileDialog
    from openpyxl import load_workbook
    root= Tkinter.Tk()
    root.withdraw()
    file_opt = options = {}
    options['defaultextension'] = default_extension
    options['filetypes'] = possible_extensions
    options['initialdir'] = 'D:\\programmation\workspace\eclipse\maitrise'
    options['initialfile'] = generic_file_name
    options['parent'] = root
    options['title'] = string_file_to_open
    file_path = tkFileDialog.askopenfilename(**file_opt)

    wb = load_workbook(filename = file_path, use_iterators = True, data_only=True)
    ws = wb.get_active_sheet()
    i=0
    
    dict_data = {}
    headers = ["CAS","ED_50_inh_non_canc_lifetime_kg","ED_50_ing_non_canc_lifetime_kg","ED_50_inh_canc_lifetime_kg","ED_50_ing_canc_lifetime_kg","GSD2_cancer","GSD2_non_cancer","Flag_cancer_inh","Flag_cancer_ing","Flag_non_cancer_inh","Flag_non_cancer_ing"]
    first = 0
    for row in ws.iter_rows():
        count = 0
        if first == 0:
            first = 1
            continue
        else:
            data = {}
            for cell in row:
                if cell.value == "n/a" or cell.value == "NEG":
                    data[headers[count]] = 0
                else:
                    data[headers[count]] = cell.value
                count += 1
            i+=1
        if dict_data.setdefault(data["CAS"], {}) == {}:
            dict_data[data["CAS"]] = data
        else:
            print "error, duplicate line if USEtox ED_50 file for CAS " + data["CAS"]
            print dict_data
            quit()
    data=[dict_data,ws.get_highest_row(),ws.get_highest_column()]
    return data;

#for import_severity the input file is an excel file with  one ligne header:
#effect						severity Daly per case	GSD2
#"cancer" or "non-cancer"	float					float (usetox: 2,7 for cancer, 13 for non-cancer)
#outputs a list: [{"cancer" : ["Effect","Severity","GSD2"] , "non-cancer" : ["Effect","Severity","GSD2"]},nb_row of initial fila, nb column initial file]

def import_severity(string_file_to_open,generic_file_name,default_extension,possible_extensions):
    import Tkinter, tkFileDialog
    from openpyxl import load_workbook
    root= Tkinter.Tk()
    root.withdraw()
    file_opt = options = {}
    options['defaultextension'] = default_extension
    options['filetypes'] = possible_extensions
    options['initialdir'] = 'D:\\programmation\workspace\eclipse\maitrise'
    options['initialfile'] = generic_file_name
    options['parent'] = root
    options['title'] = string_file_to_open
    file_path = tkFileDialog.askopenfilename(**file_opt)

    wb = load_workbook(filename = file_path, use_iterators = True, data_only=True)
    ws = wb.get_active_sheet()
    i=0
    
    dict_data = {}
    headers = ["Effect","Severity","GSD2"]
    first = 1
    for row in ws.iter_rows():
        count = 0
        if first == 1:
            first = 0
            continue
        else:
            data = {}
            for cell in row:
                data[headers[count]] = cell.value
                count += 1
            i+=1
        if dict_data.setdefault(data["Effect"], {}) == {}:
            dict_data[data["Effect"]] = data
        else:
            print "error, duplicate line if USEtox ED_50 file for CAS " + data["CAS"]
            print dict_data
            quit()
    data=[dict_data,ws.get_highest_row(),ws.get_highest_column()]
    return data;